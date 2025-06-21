# main.py

import sys
import os
import datetime
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QHBoxLayout, QVBoxLayout, QLabel,
    QLineEdit, QPushButton, QListWidget, QListWidgetItem, QCalendarWidget,
    QScrollArea, QCheckBox, QToolTip, QDialog, QFormLayout, QTextEdit,
    QDateEdit, QDialogButtonBox, QMenu, QFrame, QMessageBox, QDateTimeEdit,
    QFileDialog
)
from PyQt6.QtGui import (
    QIcon, QFont, QPalette, QColor, QPainter, QCursor
)
from PyQt6.QtCore import (
    Qt, QSize, pyqtSignal, QDate, QPropertyAnimation, QEasingCurve, QDateTime,
    QParallelAnimationGroup, QAbstractAnimation, QPoint, QTimer
)

from database import DatabaseManager

# --- Зависимость для экспорта в Excel ---
try:
    import openpyxl # type: ignore
    from openpyxl.styles import Font, Alignment # type: ignore
    from openpyxl.utils import get_column_letter # type: ignore
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# --- Вспомогательные функции ---

def clear_layout(layout):
    """Рекурсивно очищает layout от всех виджетов."""
    if layout is not None:
        while layout.count():
            item = layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()
            else:
                clear_layout(item.layout())

def load_icon(icon_path):
    """Безопасно загружает иконку по пути."""
    if os.path.exists(icon_path):
        return QIcon(icon_path)
    print(f"Внимание: Иконка не найдена по пути {icon_path}")
    return QIcon()

def colorize_icon(icon: QIcon, color: QColor) -> QIcon:
    """Перекрашивает иконку в заданный цвет (для соответствия теме)."""
    if icon.isNull():
        return icon
    pixmap = icon.pixmap(QSize(16, 16))
    painter = QPainter(pixmap)
    painter.setCompositionMode(QPainter.CompositionMode.CompositionMode_SourceIn)
    painter.fillRect(pixmap.rect(), color)
    painter.end()
    return QIcon(pixmap)

# --- Классы виджетов ---

class ClickableLabel(QLabel):
    """Метка, которая реагирует на клики."""
    clicked = pyqtSignal()
    def __init__(self, text, parent=None):
        super().__init__(text, parent)
        self.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
    def mousePressEvent(self, event):
        self.clicked.emit()
        super().mousePressEvent(event)

class AboutDialog(QDialog):
    """Диалоговое окно "О приложении"."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("О приложении")
        self.setFixedSize(350, 200)
        layout = QVBoxLayout(self)
        layout.setAlignment(Qt.AlignmentFlag.AlignHCenter)
        title_label = QLabel("Zettelkasten")
        title_label.setFont(QFont("Inter", 22, QFont.Weight.Bold))
        title_label.setAlignment(Qt.AlignmentFlag.AlignHCenter)
        version_label = QLabel("Версия 0.1a")
        version_label.setAlignment(Qt.AlignmentFlag.AlignHCenter)
        separator = QFrame()
        separator.setFrameShape(QFrame.Shape.HLine)
        separator.setFrameShadow(QFrame.Shadow.Sunken)

        help_text = """
        <p><b>Zettelkasten</b> — это ваш персональный менеджер задач, 
        созданный для того, чтобы помочь вам организовать свои дела 
        и ничего не забыть.</p>
        
        <p><b>Добавление задач:</b> Нажмите кнопку <i>"+ Новая задача"</i> вверху,
        чтобы создать новую запись. Вы можете сразу добавить ее в "Важное" или "Личное".</p>
        
        <p><b>Редактирование:</b> <b>Дважды кликните</b> по любой задаче, 
        чтобы изменить ее название, добавить детали, теги, срок выполнения или 
        настроить напоминания.</p>
        
        <p><b>Завершение:</b> Поставьте <b>галочку</b> слева от задачи, 
        чтобы отметить ее как выполненную.</p>
        
        <p><b>Важность:</b> Нажмите на <b>звездочку (☆)</b>, чтобы сделать задачу важной.</p>
        
        <p><b>Навигация:</b> Используйте меню слева для быстрой фильтрации задач.</p>
        
        <p><b>Календарь:</b> Кликните на любую дату в календаре справа, 
        чтобы увидеть все задачи с этим сроком выполнения.</p>
        
        <p><b>Отчеты:</b> Нажмите кнопку <i>"Выгрузить отчет"</i>, чтобы сохранить 
        список задач за выбранный период.</p>
        """
        help_label = QLabel(help_text)
        help_label.setWordWrap(True) 

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
        buttons.accepted.connect(self.accept)

        layout.addWidget(title_label)
        layout.addWidget(version_label)
        layout.addWidget(separator)
        layout.addWidget(help_label) # Просто добавляем label в layout
        layout.addStretch() # Добавляем растягивающееся пространство, чтобы кнопки были внизу
        layout.addWidget(buttons)

class AddTaskDialog(QDialog):
    """Диалог для добавления новой задачи."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Добавить новую задачу")
        self.setMinimumWidth(400)
        self.layout = QVBoxLayout(self)
        form_layout = QFormLayout()
        self.title_edit = QLineEdit()
        self.details_edit = QTextEdit()
        self.details_edit.setAcceptRichText(False)
        self.tags_edit = QLineEdit()
        self.tags_edit.setPlaceholderText("Например: Работа, Личное, Дом")
        self.due_date_edit = QDateEdit(self)
        self.due_date_edit.setCalendarPopup(True)
        self.due_date_edit.setDate(QDate.currentDate())
        self.important_check = QCheckBox("Отметить как важное")
        form_layout.addRow("Название:", self.title_edit)
        form_layout.addRow("Детали:", self.details_edit)
        form_layout.addRow("Теги (через запятую):", self.tags_edit)
        form_layout.addRow("Срок выполнения:", self.due_date_edit)
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        self.layout.addLayout(form_layout)
        self.layout.addWidget(self.important_check)
        self.layout.addWidget(button_box)

    def get_task_data(self):
        """Собирает данные из полей формы в словарь."""
        return {"title": self.title_edit.text().strip(),
                "details": self.details_edit.toPlainText().strip(),
                "tags": self.tags_edit.text().strip(),
                "due_date": self.due_date_edit.date().toPyDate().isoformat(),
                "is_important": self.important_check.isChecked()}

class EditTaskDialog(QDialog):
    """Диалог для редактирования существующей задачи и ее напоминаний."""
    def __init__(self, task_data, reminders, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Редактировать задачу")
        self.setMinimumWidth(450)
        self.layout = QVBoxLayout(self)
        form_layout = QFormLayout()
        self.title_edit = QLineEdit()
        self.details_edit = QTextEdit()
        self.details_edit.setAcceptRichText(False)
        self.tags_edit = QLineEdit()
        self.due_date_edit = QDateEdit(self)
        self.due_date_edit.setCalendarPopup(True)
        self.important_check = QCheckBox("Отметить как важное")
        form_layout.addRow("Название:", self.title_edit)
        form_layout.addRow("Детали:", self.details_edit)
        form_layout.addRow("Теги (через запятую):", self.tags_edit)
        form_layout.addRow("Срок выполнения:", self.due_date_edit)
        self.layout.addLayout(form_layout)
        self.layout.addWidget(self.important_check)

        # Разделитель
        separator = QFrame()
        separator.setFrameShape(QFrame.Shape.HLine)
        separator.setFrameShadow(QFrame.Shadow.Sunken)
        self.layout.addWidget(separator)

        # Секция напоминаний
        self.layout.addWidget(QLabel("<b>Напоминания</b>"))
        self.reminders_list = QListWidget()
        self.reminders_list.setToolTip("Двойной клик для удаления напоминания.")
        self.reminders_list.itemDoubleClicked.connect(self.remove_selected_reminder)
        self.layout.addWidget(self.reminders_list)
        
        # Контролы для добавления напоминаний
        reminder_controls_layout = QHBoxLayout()
        self.reminder_datetime_edit = QDateTimeEdit(self)
        self.reminder_datetime_edit.setCalendarPopup(True)
        self.reminder_datetime_edit.setDateTime(QDateTime.currentDateTime().addSecs(3600))
        self.reminder_datetime_edit.setDisplayFormat("dd.MM.yyyy HH:mm")
        add_reminder_btn = QPushButton("Добавить")
        add_reminder_btn.clicked.connect(self.add_reminder_to_list)
        reminder_controls_layout.addWidget(self.reminder_datetime_edit, 1)
        reminder_controls_layout.addWidget(add_reminder_btn)
        self.layout.addLayout(reminder_controls_layout)

        # Кнопки OK/Cancel
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        self.layout.addWidget(button_box)
        self.populate_data(task_data, reminders)

    def populate_data(self, task_data, reminders):
        """Заполняет поля формы данными существующей задачи."""
        self.title_edit.setText(task_data.get('title', ''))
        self.details_edit.setText(task_data.get('details', ''))
        self.tags_edit.setText(task_data.get('tags', ''))
        if due_date_str := task_data.get('due_date'):
            self.due_date_edit.setDate(QDate.fromString(due_date_str, "yyyy-MM-dd"))
        self.important_check.setChecked(bool(task_data.get('is_important', 0)))
        
        for reminder in reminders:
            dt = QDateTime.fromString(reminder['reminder_datetime'], Qt.DateFormat.ISODate)
            item = QListWidgetItem(dt.toString("dd MMMM yy 'в' HH:mm"))
            item.setData(Qt.ItemDataRole.UserRole, reminder['reminder_datetime'])
            self.reminders_list.addItem(item)

    def add_reminder_to_list(self):
        """Добавляет новое напоминание в список (без дубликатов)."""
        dt = self.reminder_datetime_edit.dateTime()
        iso_string = dt.toString(Qt.DateFormat.ISODate)
        
        for i in range(self.reminders_list.count()):
            if self.reminders_list.item(i).data(Qt.ItemDataRole.UserRole) == iso_string:
                return  # Не добавлять, если уже существует
                
        item = QListWidgetItem(dt.toString("dd MMMM yy 'в' HH:mm"))
        item.setData(Qt.ItemDataRole.UserRole, iso_string)
        self.reminders_list.addItem(item)
        self.reminders_list.sortItems()

    def remove_selected_reminder(self, item):
        """Удаляет выбранное напоминание из списка."""
        self.reminders_list.takeItem(self.reminders_list.row(item))

    def get_task_data(self):
        """Собирает данные о задаче из полей формы."""
        return {"title": self.title_edit.text().strip(),
                "details": self.details_edit.toPlainText().strip(),
                "tags": self.tags_edit.text().strip(),
                "due_date": self.due_date_edit.date().toPyDate().isoformat(),
                "is_important": self.important_check.isChecked()}

    def get_reminders_data(self):
        """Собирает список всех напоминаний из виджета."""
        return [self.reminders_list.item(i).data(Qt.ItemDataRole.UserRole) for i in range(self.reminders_list.count())]

class ReportDialog(QDialog):
    """Диалог для выбора диапазона дат для отчета."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Выгрузить отчет по задачам")
        self.setMinimumWidth(350)
        self.layout = QVBoxLayout(self)
        form_layout = QFormLayout()
        self.start_date_edit = QDateEdit(self, calendarPopup=True, date=QDate.currentDate().addDays(-7))
        self.end_date_edit = QDateEdit(self, calendarPopup=True, date=QDate.currentDate())
        form_layout.addRow("Начальная дата:", self.start_date_edit)
        form_layout.addRow("Конечная дата:", self.end_date_edit)
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        self.layout.addLayout(form_layout)
        self.layout.addWidget(button_box)

    def get_date_range(self):
        """Возвращает выбранный диапазон дат."""
        return {"start_date": self.start_date_edit.date().toPyDate().isoformat(), 
                "end_date": self.end_date_edit.date().toPyDate().isoformat()}

class TaskWidget(QWidget):
    """Виджет для отображения одной задачи в списке."""
    status_changed = pyqtSignal(int, bool)
    importance_changed = pyqtSignal(int, bool)
    edit_requested = pyqtSignal(int)
    
    def __init__(self, task_data):
        super().__init__()
        self.task_id = task_data['id']
        self.setObjectName("TaskWidget")
        # Начальные значения для анимации появления
        self.setWindowOpacity(0.0)
        self.setMaximumHeight(0)

        # --- Создание layout и виджетов ---
        layout = QHBoxLayout(self)
        layout.setContentsMargins(10, 5, 10, 5)
        layout.setSpacing(15)
        
        self.checkbox = QCheckBox()
        self.checkbox.setChecked(bool(task_data['is_completed']))
        self.checkbox.stateChanged.connect(self.on_status_change)
        
        # Текстовая часть (Название и мета-информация)
        text_layout = QVBoxLayout()
        text_layout.setSpacing(2)
        title_label = QLabel(task_data['title'])
        title_label.setObjectName("TaskTitle")
        title_label.setWordWrap(True)
        
        meta_text = []
        if task_data['tags']: meta_text.append(task_data['tags'])
        if task_data['due_date']:
            try: meta_text.append(datetime.date.fromisoformat(task_data['due_date']).strftime("%b %d"))
            except (ValueError, TypeError): pass
            
        meta_label = QLabel(" • ".join(meta_text))
        meta_label.setObjectName("TaskMeta")
        text_layout.addWidget(title_label)
        if meta_text: text_layout.addWidget(meta_label)
        
        # Кнопка "Важное"
        self.star_button = QPushButton()
        self.star_button.setObjectName("StarButton")
        self.star_button.setFixedSize(30, 30)
        self.star_button.setCheckable(True)
        self.star_button.setChecked(bool(task_data['is_important']))
        self.star_button.setText("★" if self.star_button.isChecked() else "☆")
        self.star_button.clicked.connect(self.on_importance_change)
        
        # Сборка layout
        layout.addWidget(self.checkbox)
        layout.addLayout(text_layout)
        layout.addStretch()
        layout.addWidget(self.star_button)
        if task_data['details']: self.setToolTip(f"<b>Детали:</b><br>{task_data['details']}")
        
        self.update_visual_state(bool(task_data['is_completed']))

    def update_visual_state(self, is_completed):
        """Обновляет стиль виджета в зависимости от статуса (выполнено/не выполнено)."""
        self.setProperty("completed", is_completed)
        self.style().unpolish(self)
        self.style().polish(self)

    def on_status_change(self, state):
        """Сигнал при изменении состояния чекбокса."""
        is_completed = (state == Qt.CheckState.Checked.value)
        self.update_visual_state(is_completed)
        self.status_changed.emit(self.task_id, is_completed)

    def on_importance_change(self):
        """Сигнал при нажатии на кнопку 'важное'."""
        is_important = self.star_button.isChecked()
        self.star_button.setText("★" if is_important else "☆")
        self.importance_changed.emit(self.task_id, is_important)

    def mouseDoubleClickEvent(self, event):
        """Сигнал для редактирования по двойному клику."""
        self.edit_requested.emit(self.task_id)
        super().mouseDoubleClickEvent(event)


# --- Главное окно приложения ---
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.db = DatabaseManager()
        self.current_filter = 'important'
        self.current_filter_value = None
        self.current_title = "Важное"
        self.active_animations = []
        
        self.setWindowTitle("Zettelkasten")
        self.setGeometry(100, 100, 1280, 800)
        
        # Загрузка и раскрашивание иконок
        text_color = self.palette().color(QPalette.ColorRole.Text)
        raw_icons = {
            "important": load_icon("icons/important.svg"),
            "personal": load_icon("icons/personal.svg"),
            "completed": load_icon("icons/completed.svg"),
            "tag": load_icon("icons/tag.svg"),
        }
        self.icons = {name: colorize_icon(icon, text_color) for name, icon in raw_icons.items()}
        
        # Инициализация UI
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QHBoxLayout(main_widget)
        main_layout.setSpacing(0)
        main_layout.setContentsMargins(0, 0, 0, 0)
        self.init_ui(main_layout)
        
        # Первоначальное обновление данных
        self.refresh_all_views(animated=True)

        # Таймер для проверки напоминаний каждые 30 секунд
        self.reminder_timer = QTimer(self)
        self.reminder_timer.timeout.connect(self.check_for_reminders)
        self.reminder_timer.start(30000)

    # --- Инициализация и настройка UI ---
    
    def init_ui(self, main_layout):
        """Инициализирует основной интерфейс, разделенный на три панели."""
        left_panel = self.create_left_panel()
        center_panel = self.create_center_panel()
        right_panel = self.create_right_panel()
        main_layout.addWidget(left_panel)
        main_layout.addWidget(center_panel, 1) # центральная панель растягивается
        main_layout.addWidget(right_panel)

    def create_left_panel(self):
        """Создает левую панель с навигацией (поиск, избранное, теги)."""
        left_panel = QWidget()
        left_panel.setObjectName("LeftPanel")
        left_panel.setFixedWidth(250)
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(10, 10, 10, 10)
        left_layout.setSpacing(10)
        
        title_label = QLabel("Zettelkasten")
        title_label.setObjectName("AppTitle")
        
        self.search_bar = QLineEdit(placeholderText="🔍 Поиск")
        self.search_bar.setObjectName("SearchBar")
        self.search_bar.textChanged.connect(self.on_search_text_changed)
        
        self.favorites_list = QListWidget()
        self.favorites_list.setObjectName("NavList")
        self.favorites_list.itemClicked.connect(self.on_nav_item_clicked)
        
        self.tags_list = QListWidget()
        self.tags_list.setObjectName("NavList")
        self.tags_list.itemClicked.connect(self.on_tag_item_clicked)
        
        self.report_button = QPushButton("Выгрузить отчет")
        self.report_button.setObjectName("ReportButton")
        self.report_button.clicked.connect(self.show_report_dialog)

        left_layout.addWidget(title_label)
        left_layout.addWidget(self.search_bar)
        left_layout.addSpacing(10)
        left_layout.addWidget(QLabel("Избранное"))
        left_layout.addWidget(self.favorites_list)
        left_layout.addSpacing(10)
        left_layout.addWidget(QLabel("Ваши теги"))
        left_layout.addWidget(self.tags_list, 1)
        left_layout.addWidget(self.report_button)
        return left_panel

    def create_center_panel(self):
        """Создает центральную панель для отображения списка задач."""
        center_panel = QWidget()
        center_panel.setObjectName("CenterPanel")
        center_layout = QVBoxLayout(center_panel)
        center_layout.setContentsMargins(20, 20, 20, 20)
        
        header_layout = QHBoxLayout()
        self.center_title_label = QLabel(self.current_title)
        self.center_title_label.setObjectName("CenterTitle")
        new_task_button = QPushButton("+ Новая задача")
        new_task_button.setObjectName("NewTaskButton")
        new_task_button.clicked.connect(self.show_new_task_menu)
        header_layout.addWidget(self.center_title_label)
        header_layout.addStretch()
        header_layout.addWidget(new_task_button)
        
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setObjectName("ScrollArea")
        tasks_container = QWidget()
        self.tasks_layout = QVBoxLayout(tasks_container)
        self.tasks_layout.setSpacing(5)
        self.tasks_layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        scroll_area.setWidget(tasks_container)
        
        center_layout.addLayout(header_layout)
        center_layout.addWidget(scroll_area)
        return center_panel

    def create_right_panel(self):
        """Создает правую панель с календарем и списком завершенных задач."""
        right_panel = QWidget()
        right_panel.setObjectName("RightPanel")
        right_panel.setFixedWidth(300)
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(10, 15, 10, 15)
        
        profile_label = ClickableLabel("Справка")
        profile_label.setFont(QFont("Inter", 12, QFont.Weight.Bold))
        profile_label.clicked.connect(self.show_about_dialog)
        
        self.calendar = QCalendarWidget(verticalHeaderFormat=QCalendarWidget.VerticalHeaderFormat.NoVerticalHeader, gridVisible=True)
        self.calendar.setObjectName("CalendarWidget")
        self.calendar.selectionChanged.connect(self.on_date_selected)
        
        completed_label = QLabel("Завершенные задачи")
        completed_label.setFont(QFont("Inter", 12, QFont.Weight.Bold))
        self.completed_list_widget = QListWidget()
        self.completed_list_widget.setObjectName("CompletedList")
        
        right_layout.addWidget(profile_label, alignment=Qt.AlignmentFlag.AlignTop)
        right_layout.addSpacing(20)
        right_layout.addWidget(self.calendar)
        right_layout.addSpacing(20)
        right_layout.addWidget(completed_label)
        right_layout.addWidget(self.completed_list_widget)
        return right_panel

    # --- Напоминания и Анимации ---

    def check_for_reminders(self):
        """Проверяет и отображает напоминания, срок которых наступил."""
        now_iso = datetime.datetime.now().isoformat()
        due_reminders = self.db.get_due_reminders(now_iso)

        for reminder in due_reminders:
            msg_box = QMessageBox(self, icon=QMessageBox.Icon.Information, windowTitle="Напоминание о задаче")
            dt = QDateTime.fromString(reminder['reminder_datetime'], Qt.DateFormat.ISODate)
            msg_box.setText(f"<b>{reminder['title']}</b>")
            msg_box.setInformativeText(f"Время выполнить задачу! (Напоминание на {dt.toString('HH:mm')})")
            msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
            msg_box.exec()
            self.db.delete_reminder(reminder['reminder_id'])

    def animate_show_item(self, widget, duration):
        """Анимация плавного появления виджета (изменение высоты и прозрачности)."""
        group = QParallelAnimationGroup(self)
        opacity_anim = QPropertyAnimation(widget, b"windowOpacity")
        opacity_anim.setDuration(duration)
        opacity_anim.setStartValue(0.0)
        opacity_anim.setEndValue(1.0)
        size_anim = QPropertyAnimation(widget, b"maximumHeight")
        size_anim.setDuration(duration)
        size_anim.setStartValue(0)
        size_anim.setEndValue(widget.sizeHint().height())
        size_anim.setEasingCurve(QEasingCurve.Type.OutCubic)
        group.addAnimation(opacity_anim)
        group.addAnimation(size_anim)
        group.finished.connect(lambda: self.active_animations.remove(group) if group in self.active_animations else None)
        self.active_animations.append(group)
        group.start(QAbstractAnimation.DeletionPolicy.DeleteWhenStopped)

    # --- Обновление данных в UI ---

    def refresh_all_views(self, animated=False):
        """Обновляет все списки и панели в приложении."""
        self.refresh_left_panel()
        self.refresh_task_list(animated)
        self.refresh_completed_list()

    def refresh_task_list(self, animated=False, tasks_list=None):
        """Обновляет центральный список задач в соответствии с текущим фильтром."""
        clear_layout(self.tasks_layout)
        tasks = tasks_list if tasks_list is not None else self.db.get_tasks(filter_by=self.current_filter, value=self.current_filter_value)
        for i, task_data in enumerate(tasks):
            task_widget = TaskWidget(task_data)
            task_widget.status_changed.connect(self.handle_task_status_change)
            task_widget.importance_changed.connect(self.handle_task_importance_change)
            task_widget.edit_requested.connect(self.show_edit_task_dialog)
            self.tasks_layout.addWidget(task_widget)
            if animated:
                self.animate_show_item(task_widget, 250 + i * 25)

    def refresh_left_panel(self):
        """Обновляет списки 'Избранное' и 'Теги' в левой панели."""
        # Обновление "Избранного"
        self.favorites_list.clear()
        self.favorites_list.addItem(QListWidgetItem(self.icons.get("important"), "Важное"))
        self.favorites_list.addItem(QListWidgetItem(self.icons.get("personal"), "Личное"))
        self.favorites_list.addItem(QListWidgetItem(self.icons.get("completed"), "Завершенные"))

        # Обновление списка тегов со счетчиками
        self.tags_list.clear()
        for tag, count in sorted(self.db.get_tags_with_counts().items()):
            item = QListWidgetItem(self.tags_list)
            # Создаем кастомный виджет для строки тега
            row_widget = QWidget()
            row_layout = QHBoxLayout(row_widget)
            row_layout.setContentsMargins(5, 3, 8, 3) 
            row_layout.setSpacing(6)
            icon_label = QLabel()
            icon_label.setPixmap(self.icons.get("tag").pixmap(QSize(16, 16)))
            count_label = QLabel(str(count))
            count_label.setObjectName("TagCount")
            count_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            row_layout.addWidget(icon_label)
            row_layout.addWidget(QLabel(tag), 1)
            row_layout.addWidget(count_label)
            item.setData(Qt.ItemDataRole.UserRole, tag) # Сохраняем имя тега для обработчика
            self.tags_list.setItemWidget(item, row_widget)
            
    def refresh_completed_list(self):
        """Обновляет список последних завершенных задач в правой панели."""
        self.completed_list_widget.clear()
        for task in self.db.get_tasks(filter_by='completed')[:5]:
            item = QListWidgetItem(f"✔ {task['title']}")
            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsSelectable) # Делаем невыделяемым
            self.completed_list_widget.addItem(item)

    # --- Отображение диалоговых окон ---

    def show_about_dialog(self):
        """Показывает диалог 'О приложении'."""
        AboutDialog(self).exec()

    def show_new_task_menu(self):
        """Показывает контекстное меню для кнопки 'Новая задача'."""
        button = self.sender()
        menu = QMenu(self)
        menu.addAction("Добавить как важное", lambda: self.show_add_task_dialog(mark_as_important=True))
        menu.addAction("Добавить в 'Личное'", lambda: self.show_add_task_dialog(add_tag='Личное'))
        menu.exec(button.mapToGlobal(QPoint(0, button.height())))

    def show_add_task_dialog(self, mark_as_important=False, add_tag=None):
        """Открывает диалог добавления задачи и обрабатывает результат."""
        dialog = AddTaskDialog(self)
        if mark_as_important: dialog.important_check.setChecked(True)
        if add_tag: dialog.tags_edit.setText(add_tag)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            task_data = dialog.get_task_data()
            if task_data['title']: # Добавляем задачу только если есть заголовок
                self.db.add_task(**task_data)
                self.refresh_all_views(animated=True)

    def show_edit_task_dialog(self, task_id):
        """Открывает диалог редактирования задачи и обрабатывает результат."""
        task_data = self.db.get_task_by_id(task_id)
        if not task_data: return
        reminders = self.db.get_reminders_for_task(task_id)
        dialog = EditTaskDialog(task_data, reminders, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            new_data = dialog.get_task_data()
            if new_data['title']:
                self.db.update_task(task_id, new_data)
                self.db.replace_all_reminders_for_task(task_id, dialog.get_reminders_data())
                self.refresh_all_views(animated=True)

    # --- Обработчики событий от виджетов ---

    def handle_task_status_change(self, task_id, is_completed):
        """Обрабатывает изменение статуса задачи (выполнена/не выполнена)."""
        self.db.update_task_status(task_id, is_completed)
        # Удаляем виджет из списка активных задач
        for i in range(self.tasks_layout.count()):
            widget = self.tasks_layout.itemAt(i).widget()
            if isinstance(widget, TaskWidget) and widget.task_id == task_id:
                widget.deleteLater()
                break
        # Обновляем списки, где это изменение должно отразиться
        self.refresh_completed_list()
        self.refresh_left_panel()

    def handle_task_importance_change(self, task_id, is_important):
        """Обрабатывает изменение флага 'важное' у задачи."""
        self.db.update_task_importance(task_id, is_important)
        # Если мы находимся в фильтре "Важное", список нужно перерисовать
        if self.current_filter == 'important':
             self.refresh_task_list(animated=True)

    # --- Обработчики навигации и поиска ---

    def on_nav_item_clicked(self, item):
        """Обрабатывает клик по элементам в списке 'Избранное'."""
        self.search_bar.clear() # Очищаем поиск
        filter_text = item.text()
        self.current_title = filter_text
        self.current_filter_value = None
        if filter_text == "Важное": self.current_filter = 'important'
        elif filter_text == "Завершенные": self.current_filter = 'completed'
        elif filter_text == "Личное": self.current_filter, self.current_filter_value = 'tag', 'Личное'
        self.center_title_label.setText(self.current_title)
        self.refresh_task_list(animated=True)

    def on_tag_item_clicked(self, item):
        """Обрабатывает клик по тегу в списке тегов."""
        self.search_bar.clear()
        if tag_name := item.data(Qt.ItemDataRole.UserRole):
            self.current_filter = 'tag'
            self.current_filter_value = tag_name
            self.current_title = f"Тег: {tag_name}"
            self.center_title_label.setText(self.current_title)
            self.refresh_task_list(animated=True)

    def on_date_selected(self):
        """Обрабатывает выбор даты в календаре."""
        self.search_bar.clear()
        selected_date_q = self.calendar.selectedDate()
        self.current_filter = 'date'
        self.current_filter_value = selected_date_q.toString("yyyy-MM-dd")
        self.current_title = f"Задачи на {selected_date_q.toString('d MMMM yyyy г.')}"
        self.center_title_label.setText(self.current_title)
        self.refresh_task_list(animated=True)

    def on_search_text_changed(self, text):
        """Обрабатывает изменение текста в строке поиска."""
        query = text.strip()
        if query:
            self.center_title_label.setText(f'Результаты поиска: "{query}"')
            self.refresh_task_list(animated=True, tasks_list=self.db.search_tasks(query))
        else: # Если поиск пуст, возвращаемся к последнему активному фильтру
            self.center_title_label.setText(self.current_title)
            self.refresh_task_list(animated=True)

    # --- Создание и сохранение отчетов ---
    
    def show_report_dialog(self):
        """Показывает диалог выбора дат и инициирует сохранение отчета."""
        dialog = ReportDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            date_range = dialog.get_date_range()
            start_iso, end_iso = date_range["start_date"], date_range["end_date"]
            
            report_tasks = self.db.get_tasks(filter_by='date_range', start_date=start_iso, end_date=end_iso)
            if not report_tasks:
                QMessageBox.information(self, "Нет данных", "Задачи не найдены за выбранный период.")
                return

            default_filename = f"Отчет по задачам {start_iso} - {end_iso}"
            filters = "Excel Files (*.xlsx);;Text Files (*.txt)"
            filePath, selected_filter = QFileDialog.getSaveFileName(self, "Сохранить отчет", default_filename, filters)

            if filePath:
                if 'Excel' in selected_filter:
                    self.save_report_as_excel(report_tasks, filePath, start_iso, end_iso)
                else:
                    self.save_report_as_txt(report_tasks, filePath, start_iso, end_iso)

    def save_report_as_txt(self, tasks, file_path, start_date, end_date):
        """Формирует и сохраняет отчет в формате .txt."""
        start_pretty = QDate.fromString(start_date, 'yyyy-MM-dd').toString('dd.MM.yyyy')
        end_pretty = QDate.fromString(end_date, 'yyyy-MM-dd').toString('dd.MM.yyyy')
        report_text = f"Отчет по задачам с {start_pretty} по {end_pretty}:\n{'=' * 40}\n\n"
        
        for task in tasks:
            status = "✔️ Выполнено" if task['is_completed'] else "❌ Не выполнено"
            due_date = f"Срок: {QDate.fromString(task['due_date'], 'yyyy-MM-dd').toString('dd.MM.yyyy')}" if task['due_date'] else "Срок не указан"
            report_text += f"Задача: {task['title']}\n"
            report_text += f"Статус: {status} | {due_date}\n"
            if task['details']: report_text += f"  Детали: {task['details']}\n"
            if task['tags']: report_text += f"  Теги: {task['tags']}\n"
            report_text += f"{'-' * 40}\n"
        
        try:
            with open(file_path, 'w', encoding='utf-8') as f: f.write(report_text)
            QMessageBox.information(self, "Успех", f"Отчет успешно сохранен в файл:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить файл отчета.\nОшибка: {e}")

    def save_report_as_excel(self, tasks, file_path, start_date, end_date):
        """Формирует и сохраняет отчет в формате .xlsx."""
        if not OPENPYXL_AVAILABLE:
            QMessageBox.critical(self, "Ошибка", "Для экспорта в Excel необходимо установить библиотеку openpyxl.\nВыполните: pip install openpyxl")
            return

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        start_pretty = QDate.fromString(start_date, 'yyyy-MM-dd').toString('dd.MM.yyyy')
        end_pretty = QDate.fromString(end_date, 'yyyy-MM-dd').toString('dd.MM.yyyy')
        sheet.title = f"Отчет {start_pretty}-{end_pretty}"
        
        headers = ["Задача", "Статус", "Срок выполнения", "Детали", "Теги"]
        sheet.append(headers)
        header_font = Font(bold=True)
        for col_num, header_title in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for task in tasks:
            status = "Выполнено" if task['is_completed'] else "Не выполнено"
            due_date = QDate.fromString(task['due_date'], 'yyyy-MM-dd').toString('dd.MM.yyyy') if task.get('due_date') else ""
            sheet.append([task.get('title', ''), status, due_date, task.get('details', ''), task.get('tags', '')])

        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = max(len(str(cell.value)) for cell in sheet[column_letter] if cell.value)
            sheet.column_dimensions[column_letter].width = min((max_length + 2) * 1.2, 70)

        try:
            workbook.save(file_path)
            QMessageBox.information(self, "Успех", f"Отчет успешно сохранен в файл:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить файл отчета.\nОшибка: {e}")

    # --- Системные события ---
    
    def closeEvent(self, event):
        """Обрабатывает закрытие окна, корректно завершая работу с БД."""
        self.db.close()
        super().closeEvent(event)

# --- Точка входа в приложение ---
if __name__ == "__main__":
    app = QApplication(sys.argv)
    QToolTip.setFont(QFont("Inter", 10))
    if os.path.exists("icons/icons.png"):
        app.setWindowIcon(QIcon("icons/icons.png"))
    
    try:
        with open("style.qss", "r", encoding="utf-8") as f:
            app.setStyleSheet(f.read())
    except FileNotFoundError:
        print("Внимание: Файл style.qss не найден.")
    
    window = MainWindow()
    window.show()
    sys.exit(app.exec())